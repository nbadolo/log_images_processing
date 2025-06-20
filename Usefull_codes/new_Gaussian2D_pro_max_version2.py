#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Created on Tue May 13 10:21:55 2025

@author: nbadolo
"""

"""
Ce code ajuste une gaussienne 2D à ce profil en utilisant curve_fit. 
La FWHM est calculée à partir de l'écart-type σ obtenu après l'ajustement de la gaussienne. 
Le profil d'intensité et l'ajustement gaussien sont tracés pour qu'on puisse
visualiser la qualité de l'ajustement.  Plus automatisé que son grand frère Gaussian2D_pro,
il gere les psf aussi, et attribut aux étoiles qui n'ont pas de psf, la moyenne des FWHM des psf.
observées dans le même filtre. Il compare ensuite le rapport des FWHM (etoile /psf) à un seuil 
pour décider sil'étoile est résolue ou pas. Il genere aussi un fichier excel avec mise en forme 
automatique avec le rapport final.
Les inputs ainsi les résultats sont dans le dossier : 
    /home/nbadolo/Bureau/Aymard/Donnees_sph/Gaussian/Output.
    
                            CODE OKAY !!!!
"""


import os
import numpy as np
import matplotlib.pyplot as plt
from astropy.io import fits
from scipy.optimize import curve_fit
import pandas as pd
from mpl_toolkits.axes_grid1 import make_axes_locatable  # Permet de créer une colorbar compacte
from AymardPack import process_fits_image as pfi # Pour l'extraction du bruit et des pixels morts et chauds
import openpyxl
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows

# === PARAMÈTRES GLOBAUX ===
PIXEL_TO_MAS = 3.4              # Conversion de pixels en milli-arcsecondes (mas)
READ_NOISE = 0.01               # Bruit de lecture estimé (utilisé pour chi² réduit)
WINDOW_PIXELS = 100              # Taille du zoom autour de l'étoile pour le fit (en pixels)

# === GAUSSIENNE 2D CIRCULAIRE ===
def circular_gaussian_2d(coords, A, x0, y0, sigma):
    """
    Fonction de gaussienne 2D circulaire (avec le même sigma dans toutes les directions).
    :param coords: tuple de meshgrid aplatis (X, Y) pour effectuer le fit
    :param A: Amplitude de la gaussienne
    :param x0, y0: Coordonnées du centre de la gaussienne
    :param sigma: Écart-type de la gaussienne
    :return: Valeurs de la gaussienne à chaque point du meshgrid
    """
    x, y = coords
    return A * np.exp(-((x - x0)**2 + (y - y0)**2) / (2 * sigma**2)).ravel()

# === FIT SUR UNE IMAGE UNIQUE ===
def process_single_frame_2d(frame, object_name, filtre, frame_index, save_path):
    """
    Cette fonction effectue un fit d'une gaussienne circulaire 2D sur une image centrée autour de son pic.
    Les résultats sont ensuite affichés avec des contours et sauvegardés dans un fichier.
    
    :param frame: Image (frame) à analyser
    :param object_name: Nom de l'étoile à analyser
    :param filtre: Filtre utilisé pour l'image
    :param frame_index: Indice de la frame pour la gestion des fichiers
    :param save_path: Chemin où les résultats seront sauvegardés
    :return: Dictionnaire contenant les résultats du fit (FWHM, chi² réduit, etc.)
    """
    # Recherche du pic de l'étoile pour centrer le zoom
    y_max, x_max = np.unravel_index(np.argmax(frame), frame.shape)
    half = WINDOW_PIXELS // 2
    y1, y2 = y_max - half, y_max + half
    x1, x2 = x_max - half, x_max + half
    sub_img = frame[y1:y2, x1:x2].astype(float)

    # Normalisation de l'image [0, 1]
    sub_img = pfi(sub_img) # extraction des mauvais pixels 
    sub_img -= np.min(sub_img)
    #print(np.max(sub_img))
    sub_img /= np.max(sub_img)

    # Création du meshgrid pour le fit
    Y, X = np.meshgrid(np.arange(sub_img.shape[0]), np.arange(sub_img.shape[1]), indexing='ij')
    initial_guess = (1.0, sub_img.shape[1] / 2, sub_img.shape[0] / 2, 5.0)

    # Essai de fit avec la fonction gaussienne
    try:
        params, _ = curve_fit(circular_gaussian_2d, (X, Y), sub_img.ravel(), p0=initial_guess)
    except Exception as e:
        print(f"⚠️ Fit échoué pour {object_name}, frame {frame_index} : {e}")
        return None

    # Calcul des paramètres du fit
    A, x0, y0, sigma = params # ce qui donne le 4 dans l'expression du chi2_red
    FWHM_pix = 2 * np.sqrt(2 * np.log(2)) * sigma
    FWHM_mas = FWHM_pix * PIXEL_TO_MAS
    #print("param =" + str(params))
    # Calcul des résidus et du chi² réduit
    fit_2d = circular_gaussian_2d((X, Y), *params).reshape(sub_img.shape)
    residuals = sub_img - fit_2d
    sigma_data = np.sqrt(sub_img + READ_NOISE**2)
    chi2_red = np.sum((residuals / sigma_data) ** 2) / (sub_img.size - 4)

    # === AFFICHAGE ===
    # extent_mas = np.array([
    #     -sub_img.shape[1] // 2, sub_img.shape[1] // 2, 
    #     -sub_img.shape[0] // 2, sub_img.shape[0] // 2
    # ]) * PIXEL_TO_MAS
    extent_mas = [
    -half * PIXEL_TO_MAS,
    half * PIXEL_TO_MAS,
    -half * PIXEL_TO_MAS,
    half * PIXEL_TO_MAS
]

    fig, ax = plt.subplots(figsize=(6, 5))
    im = ax.imshow(sub_img, origin='lower', cmap='inferno', extent=extent_mas)

    # # Contours du fit
    # contour_levels = np.linspace(np.min(fit_2d), np.max(fit_2d), 5)
    # x_shifted = (X - x0) * PIXEL_TO_MAS
    # y_shifted = (Y - y0) * PIXEL_TO_MAS
    # ax.contour(x_shifted, y_shifted, fit_2d, levels=contour_levels, colors='white', linewidths=1.2)
    
    # Contours du fit (avec échelle logarithmique pour mieux capturer les faibles signaux)
    min_level = 0.01 * np.max(fit_2d)  # Niveau minimum: 1% du maximum
    max_level = np.max(fit_2d)
    
    # Création des niveaux logarithmiques
    num_levels = 5  # Nombre de niveaux de contours
    contour_levels = np.logspace(np.log10(min_level), np.log10(max_level), num_levels)
    
    x_shifted = (X - x0) * PIXEL_TO_MAS
    y_shifted = (Y - y0) * PIXEL_TO_MAS
    
    ax.contour(x_shifted, y_shifted, fit_2d, levels=contour_levels, colors='white', linewidths=1.2)


    # Annotations texte pour le nom de l'étoile et le filtre
    ax.text(0.02, 0.95, object_name, transform=ax.transAxes,
            fontsize=12, fontweight='bold', color='white', ha='left', va='top')
    ax.text(0.02, 0.02, filtre, transform=ax.transAxes,
            fontsize=12, fontweight='bold', color='white', ha='left', va='bottom')
    
    

    # Axes et ticks
    ax.set_xlabel("Relative RA(mas)", fontsize=11, fontweight='bold')
    ax.set_ylabel("Relative Dec (mas)", fontsize=11, fontweight='bold')
    ax.tick_params(axis='both', labelsize=9, width=1.2)
    for label in ax.get_xticklabels() + ax.get_yticklabels():
        label.set_fontweight('bold')
    ax.locator_params(axis='x', nbins=5)
    ax.locator_params(axis='y', nbins=5)

    # Colorbar compacte à droite
    divider = make_axes_locatable(ax)
    cax = divider.append_axes("right", size="5%", pad=0.05)
    cbar = fig.colorbar(im, cax=cax)
    cbar.set_label('I / Imax', fontsize=10, fontweight='bold')
    # for t in cbar.ax.get_yticklabels(): # pour mettre en gras les labels de la colorbar
    #     t.set_fontweight('bold')

    # Sauvegarde de la figure
    plt.tight_layout()
    os.makedirs(save_path, exist_ok=True)
    fig_filename = f"{object_name.replace(' ', '_')}_{filtre.replace(' ', '_')}_frame{frame_index+1}.png"
    plt.savefig(os.path.join(save_path, fig_filename))
    #plt.show()
    plt.close()

    # Retourne les résultats sous forme de dictionnaire
    return {
        'Etoile': object_name,
        'Filtre': filtre,
        'FWHM_mas': round(FWHM_mas, 2),
        'Sigma_pix': round(sigma, 2),
        'Chi2_reduit': round(chi2_red, 4)
    }

# === TRAITEMENT D'UN FICHIER FITS ===
def process_fits_file(filepath, fallback_name, save_path):
    """
    Traite un fichier FITS contenant des images et effectue le fit de gaussienne pour chaque image.
    
    :param filepath: Chemin vers le fichier FITS à traiter
    :param fallback_name: Nom de l'étoile à utiliser si le fichier FITS ne contient pas de nom d'objet
    :param save_path: Chemin où les résultats seront sauvegardés
    :return: Liste des résultats des fits pour ce fichier
    """
    results = []
    try:
        with fits.open(filepath) as hdul:
            data = hdul[0].data
            header = hdul[0].header
            object_name = header.get('OBJECT', fallback_name)
            filtre1 = header.get('HIERARCH ESO INS3 OPTI5 NAME', 'filtre1')
            filtre2 = header.get('HIERARCH ESO INS3 OPTI6 NAME', 'filtre2')

            # Si l'image est en 3D et contient deux filtres
            if data.ndim == 3 and data.shape[0] == 2:
                filters = [filtre1, filtre2]
                for i in range(2):
                    result = process_single_frame_2d(data[i], object_name, filters[i], i, save_path)
                    if result:
                        results.append(result)
            else:
                print(f"⚠️ Format inattendu pour {filepath}")
    except Exception as e:
        print(f"❌ Erreur lecture {filepath} : {e}")
    return results


def process_directory(fits_root_folder, output_root_folder):
    os.makedirs(output_root_folder, exist_ok=True)
    all_results = []
    fwhm_psf_par_filtre = {}  # Moyennes FWHM des PSF par filtre

    etoile_data = []

    for star_folder in os.listdir(fits_root_folder):
        star_path = os.path.join(fits_root_folder, star_folder)
        intensity_star_path = os.path.join(star_path, "Intensity", "star")
        intensity_psf_path = os.path.join(star_path, "Intensity", "psf")

        if not os.path.isdir(intensity_star_path):
            continue

        fits_files_star = [f for f in os.listdir(intensity_star_path) if f.endswith(".fits")]
        fits_files_psf = []
        if os.path.isdir(intensity_psf_path):
            fits_files_psf = [f for f in os.listdir(intensity_psf_path) if f.endswith(".fits")]

        if not fits_files_star:
            continue

        color = "\033[92m" if fits_files_psf else "\033[91m"
        print(f"{color}📁 Traitement de : {star_folder} {'(PSF dispo)' if fits_files_psf else '(PSF absente)'}\033[0m")

        star_fits_path = os.path.join(intensity_star_path, fits_files_star[0])
        star_results = process_fits_file(star_fits_path, fallback_name=star_folder, save_path=out_path_plot)

        if fits_files_psf:
            psf_fits_path = os.path.join(intensity_psf_path, fits_files_psf[0])
            psf_results = process_fits_file(psf_fits_path, fallback_name=star_folder + "_psf", save_path=out_path_plot)

            for s, p in zip(star_results, psf_results):
                # Enregistrer FWHM des PSF pour la moyenne
                filtre = p['Filtre']
                if filtre not in fwhm_psf_par_filtre:
                    fwhm_psf_par_filtre[filtre] = []
                fwhm_psf_par_filtre[filtre].append(p['FWHM_mas'])

                etoile_data.append({
                    **s,
                    'FWHM_psf': p['FWHM_mas'],
                    'PSF_existe': 'Yes'
                })
        else:
            for s in star_results:
                filtre = s['Filtre']
                etoile_data.append({
                    **s,
                    'FWHM_psf': None,  # Rempli plus tard
                    'PSF_existe': 'No'
                })

    # Moyennes par filtre
    moyennes_par_filtre = {
        f: round(np.mean(fwhm_list), 2) for f, fwhm_list in fwhm_psf_par_filtre.items()
    }

    # Attribution des FWHM PSF manquantes
    for entry in etoile_data:
        if entry['PSF_existe'] == 'No':
            filtre = entry['Filtre']
            fwhm_moyenne = moyennes_par_filtre.get(filtre, '-')
            entry['FWHM_psf'] = fwhm_moyenne

    # Calcul du rapport et statut résolu
    for entry in etoile_data:
        try:
            fwhm_star = float(entry['FWHM_mas'])
            fwhm_psf = float(entry['FWHM_psf'])
            ratio = round(fwhm_star / fwhm_psf, 2)
            entry['Rapport_FWHM'] = ratio
            entry['Resolution'] = 'Résolue' if ratio > 1.1 else 'Non résolue'
        except:
            entry['Rapport_FWHM'] = '-'
            entry['Resolution'] = '-'

    # Export
    # df = pd.DataFrame(etoile_data)
    
    # df.to_csv(csv_path, index=False)
    
    # Créer le DataFrame
    df = pd.DataFrame(etoile_data)
    
    # Compter résolues / non résolues
    nb_resolues = df[df['Resolution'] == 'Résolue'].shape[0]
    nb_non_resolues = df[df['Resolution'] == 'Non résolue'].shape[0]
    
    # ✅ Affichage console
    print(f"\n✅ Résumé de la résolution :")
    print(f"🔬 Nombre total d'étoiles résolues     : {nb_resolues}")
    print(f"🌑 Nombre total d'étoiles non résolues : {nb_non_resolues}")
    
    # 📁 Sauvegarde CSV et excel
    csv_path = os.path.join(output_csv, 'resultats_fwhm.csv')
    df.to_csv(csv_path, index=False)
    os.makedirs(output_csv, exist_ok=True)
    xlsx_path = os.path.join(output_csv, 'resultats_fwhm.xlsx')
    
    # 💾 Exporter vers Excel avec résumé à la fin
    with pd.ExcelWriter(xlsx_path, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='FWHM_Results')
        
        # Charger la feuille et ajouter le résumé
        workbook = writer.book
        sheet = writer.sheets['FWHM_Results']
        start_row = len(df) + 2
        sheet.cell(row=start_row, column=1, value='Résumé')
        sheet.cell(row=start_row + 1, column=1, value='Étoiles résolues')
        sheet.cell(row=start_row + 1, column=2, value=nb_resolues)
        sheet.cell(row=start_row + 2, column=1, value='Étoiles non résolues')
        sheet.cell(row=start_row + 2, column=2, value=nb_non_resolues)
    

    # 📊 Création du camembert
    labels = ['Résolues', 'Non résolues']
    sizes = [nb_resolues, nb_non_resolues]
    colors = ['#66b3ff', '#ff9999']

    if any([s is None or np.isnan(s) for s in sizes]) or sum(sizes) == 0:
        print("Erreur : Impossible de tracer le camembert, aucune étoile résolue ou non résolue détectée.")
    else:
        plt.figure(figsize=(6, 5))
        plt.pie(sizes, labels=labels, colors=colors, autopct='%1.1f%%',
                startangle=90, textprops={'fontsize': 12, 'weight': 'bold'})
        plt.title('Proportion résolues vs non résolues', fontsize=14, weight='bold')
        plt.axis('equal')
        # 📁 Dossier pour image
        chart_folder = os.path.join(output_csv, "Charts")
        os.makedirs(chart_folder, exist_ok=True)
        plt.savefig(os.path.join(chart_folder, 'resolues_vs_nonresolues_pie.png'))
        #plt.show()
        plt.close()
# === TRAITEMENT D'UN DOSSIER ENTIER ===

    # Compter les étoiles résolues et non résolues
    nb_resolues = df[df['Resolution'] == 'Résolue'].shape[0]
    nb_non_resolues = df[df['Resolution'] == 'Non résolue'].shape[0]

    print(df)
    print(f"\n✅ Résumé de la résolution :")
    print(f"🔬 Nombre total d'étoiles résolues     : {nb_resolues}")
    print(f"🌑 Nombre total d'étoiles non résolues : {nb_non_resolues}")

    print(f"\n✅ Rapport final enregistré dans : {csv_path}")

    
    
    excel_path = os.path.join(output_csv, 'rapport_fwhm_resolution.xlsx')
    df.to_excel(excel_path, index=False)
    
    # # Mise en forme
    wb = openpyxl.load_workbook(excel_path)
    ws = wb.active
    
    # Style des titres
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    
    for col_num, column_title in enumerate(df.columns, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.font = header_font
        cell.fill = header_fill
        # Ajustement automatique de la largeur
        column_letter = get_column_letter(col_num)
        max_len = max((len(str(cell.value)) for cell in ws[column_letter]), default=10)
        ws.column_dimensions[column_letter].width = max_len + 2
    
    # Coloration des étoiles résolues
    for row in range(2, ws.max_row + 1):
        res = ws.cell(row=row, column=df.columns.get_loc("Resolution") + 1).value
        if res == "Résolue":
            for col in range(1, ws.max_column + 1):
                ws.cell(row=row, column=col).fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        elif res == "Non résolue":
            for col in range(1, ws.max_column + 1):
                ws.cell(row=row, column=col).fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
    
    wb.save(excel_path)
    print(f"\n📁 Fichier Excel avec mise en forme enregistré ici : {excel_path}")
    
        



# === LANCEMENT ===
#folder_name = "test1/"
folder_name = "V854_Cen/"
#folder_name = "large_log_+/"
#folder_name = "resolved_log/"
#folder_name = "all_resolved_log/"
input_path = "/home/nbadolo/Bureau/Aymard/Donnees_sph/Gaussian/Input/" + folder_name
out_path = "/home/nbadolo/Bureau/Aymard/Donnees_sph/Gaussian/Output/version2/" + folder_name
out_path_plot = out_path + "Intensity/"
output_csv = out_path + "Csv/"




process_directory(input_path, out_path)



























